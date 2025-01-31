import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk
import pandas as pd
import openpyxl

# 전역 변수
extracted_data = None  # 추출된 데이터를 저장할 변수

# 엑셀 파일 열기
def open_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    if file_path:
        try:
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            sheet_selection.set('')  # 시트 선택 초기화
            sheet_menu['menu'].delete(0, 'end')  # 기존 메뉴 항목 지우기
            
            for sheet in sheet_names:
                sheet_menu['menu'].add_command(label=sheet, command=tk._setit(sheet_selection, sheet))
            
            current_file.set(file_path)
            messagebox.showinfo("파일 선택", f"파일이 성공적으로 열렸습니다.\n{len(sheet_names)}개의 시트가 있습니다.")
        except Exception as e:
            messagebox.showerror("오류", f"파일을 열 수 없습니다: {e}")
    else:
        messagebox.showwarning("경고", "파일을 선택하지 않았습니다.")

# 월별 데이터 필터링 후 표시
def show_data():
    global extracted_data
    file_path = current_file.get()
    if not file_path:
        messagebox.showwarning("경고", "엑셀 파일을 먼저 열어주세요.")
        return
    
    sheet_name = sheet_selection.get()
    if not sheet_name:
        messagebox.showwarning("경고", "시트를 선택해주세요.")
        return
    
    month = month_selection.get()
    if not month:
        messagebox.showwarning("경고", "월을 선택해주세요.")
        return
    
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        df.columns = ['날짜', '고객사명', '담당자명', '업무내용', '특이사항', '구분', '시작시간', '종료시간', '수행시간']
        df['날짜'] = pd.to_datetime(df['날짜'], format='%Y-%m-%d', errors='coerce')
        df = df.dropna(subset=['날짜'])  # 날짜가 NaT인 행 삭제
        df_filtered = df[df['날짜'].dt.month == int(month)]
        df_filtered = df_filtered.dropna(subset=['고객사명'])  # B열이 비어있지 않은 데이터만 유지
        
        # 추출된 데이터 저장
        extracted_data = df_filtered
        
        for i in tree.get_children():
            tree.delete(i)
        for _, row in df_filtered.iterrows():
            tree.insert('', tk.END, values=list(row))
    except Exception as e:
        messagebox.showerror("오류", f"데이터를 읽을 수 없습니다: {e}")

# 새 파일에 데이터 기록
def save_to_new_file():
    global extracted_data
    extracted_data.reset_index(drop=True, inplace=True)
    
    #월
    month = extracted_data['날짜'].dt.month
    #print(month[0])
    
    #월말
    monthEnd = extracted_data['날짜'] + pd.offsets.MonthEnd(0)
    monthEnd = monthEnd.dt.date
    #print(monthEnd[0])
    
    if extracted_data is None or extracted_data.empty:
        messagebox.showwarning("경고", "기록할 데이터가 없습니다.")
        return
    
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx;*.xls")])
    
    
    #가공된 데이터를 입력
    try:

        wb = openpyxl.load_workbook(file_path)
        sheet_names = wb.sheetnames
        sheet = wb[sheet_names[1]]
        
        sheet_first = wb[sheet_names[0]]
        sheet_second = wb[sheet_names[1]]

        # 저장된 데이터가 제대로 입력되었는지 확인
        today = pd.Timestamp.today().date()
        sheet_first.cell(row=15, column=4).value = f'{today}'
        sheet_first.cell(row=2, column=2).value = f'( {month[0]} )월 개인경비 (개인카드,영수증 등) 정산서'
        # sheet 1 일자 변경
        sheet_first.cell(row=35, column=9).value = f'{monthEnd[0]}'
        sheet_first.cell(row=36, column=9).value = f'{monthEnd[0]}'
        sheet_first.cell(row=37, column=9).value = f'{monthEnd[0]}'
        sheet_first.cell(row=38, column=9).value = f'{monthEnd[0]}'
        # sheet 1 내용 변경
        sheet_first.cell(row=35, column=14).value = f'{month[0]}월 유류대'
        sheet_first.cell(row=36, column=14).value = f'{month[0]}월 통행료'
        sheet_first.cell(row=37, column=14).value = f'{month[0]}월 대중교통'
        sheet_first.cell(row=38, column=14).value = f'{month[0]}월 주차비'
        
        sheet_second.cell(row=1, column=1).value = f'( {month[0]} )월 시내교통비 신청서'
        
        # sheet 2 데이터 기입
        for index, row in extracted_data.iterrows():
            #print(f"Processing index {index}")
            #print(f"row['날짜']: {row['날짜']}, row['업무내용']: {row['업무내용']}")
            sheet_second.cell(row=4 + index, column=1).value = row['날짜']  # A열 (날짜)
            sheet_second.cell(row=4 + index, column=3).value = f"{row['고객사명']} - {row['업무내용']}" # C열 (업무내용)
        
        sheet_first.title = f"개인경비정산서 {month[0]}월"
        wb.save(file_path)
    

        messagebox.showinfo("성공","데이터가 성공적으로 입력되었습니다.")
    except Exception as e:
        messagebox.showerror("오류",f"엑셀 파일 수정 중 오류가 발생했습니다: {e}")



# 메인 GUI 생성
root = tk.Tk()
root.title("엑셀 데이터 뷰어 및 기록기")
root.geometry("900x600")

# 엑셀 파일 경로 저장용
current_file = tk.StringVar()

# 파일 열기 버튼
open_button = tk.Button(root, text="엑셀 파일 열기", command=open_file)
open_button.pack(pady=10)

# 시트 선택 메뉴
sheet_selection = tk.StringVar()
sheet_menu_label = tk.Label(root, text="시트 선택")
sheet_menu_label.pack(pady=5)
sheet_menu = tk.OptionMenu(root, sheet_selection, [])
sheet_menu.pack(pady=5)

# 월 선택 메뉴
month_selection = tk.StringVar()
month_menu_label = tk.Label(root, text="월 선택")
month_menu_label.pack(pady=5)
month_menu = tk.OptionMenu(root, month_selection, '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12')
month_menu.pack(pady=5)

# Treeview 생성
frame = ttk.Frame(root)
frame.pack(fill=tk.BOTH, expand=True)
tree = ttk.Treeview(frame, columns=['날짜', '고객사명', '담당자명', '업무내용', '특이사항', '구분', '시작시간', '종료시간', '수행시간'], show='headings')
for col in ['날짜', '고객사명', '담당자명', '업무내용', '특이사항', '구분', '시작시간', '종료시간', '수행시간']:
    tree.heading(col, text=col)
    tree.column(col, width=120)
scroll_x = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
scroll_x.pack(side=tk.BOTTOM, fill=tk.X)
scroll_y = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
scroll_y.pack(side=tk.RIGHT, fill=tk.Y)
tree.configure(xscrollcommand=scroll_x.set, yscrollcommand=scroll_y.set)
tree.pack(fill=tk.BOTH, expand=True)

# 데이터 표시 버튼
show_button = tk.Button(root, text="데이터 보기", command=show_data)
show_button.pack(pady=5)

# 다음 단계 버튼
next_button = tk.Button(root, text="새 파일에 저장", command=save_to_new_file)
next_button.pack(pady=10)

# GUI 실행
root.mainloop()
